#
# 準備「分析」裡面的資料
#
# 組建
# 日期	種類	代號	上市櫃	產業	成分股	股期	成長股
#
#


import openpyxl

my_work_book = 'growth.xlsx'

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)


def convert_to_str(input):
    if type(input) is str:
        return input.rstrip(' ').lstrip(' ').rstrip(' ')

    if type(input) is int:
        return str(input)

    return input


def load_cell(s, r, c):
    x = s.cell(row=r, column=c)
    x.value = convert_to_str(x.value)
    return x


def set_cell(s, r, c, value):
    cell = s.cell(row=r, column=c)
    cell.value = value


sheet = wb_obj['名稱代號']

# name, 上市櫃, 產業, 50/100成分股, 有股期, 成長股
stock_info = {}
for i in range(2, sheet.max_row + 1):
    symbol = load_cell(sheet, i, 1).value
    name = load_cell(sheet, i, 2).value
    market = load_cell(sheet, i, 3).value
    category = load_cell(sheet, i, 4).value
    attr = load_cell(sheet, i, 5).value
    future = load_cell(sheet, i, 6).value
    groth = load_cell(sheet, i, 7).value

    stock_info[symbol] = (name, market, category, attr, future, groth)


sheet = wb_obj['分析']

last_date = ''
last_policy = ''

for i in range(62, sheet.max_row + 1):
    date = load_cell(sheet, i, 1)
    policy = load_cell(sheet, i, 2)
    symbol = load_cell(sheet, i, 3)

    if date.value is not None:
        last_date = date
    else:
        # set_cell(sheet, i, 1, last_date.value.strftime('%Y/%m/%d'))
        set_cell(sheet, i, 1, last_date.value)

    if policy.value is not None:
        last_policy = policy
    else:
        set_cell(sheet, i, 2, last_policy.value)

    symbol = load_cell(sheet, i, 3).value

    if symbol in stock_info.keys():
        stock = stock_info[symbol]
    else:
        continue

    set_cell(sheet, i, 4, stock[0])
    set_cell(sheet, i, 5, stock[1])
    set_cell(sheet, i, 6, stock[2])
    set_cell(sheet, i, 7, stock[3])
    set_cell(sheet, i, 8, stock[4])
    set_cell(sheet, i, 9, stock[5])

wb_obj.save(my_work_book)

exit(0)
