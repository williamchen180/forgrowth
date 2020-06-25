#
# 準備「名稱代號」裡面的資料
#
#


import openpyxl

my_work_book = 'growth.xlsx'

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)


def convert_to_str(input):
    if type(input) is str:
        return input.rstrip(' ').lstrip(' ').rstrip(' ')

    if type(input) is int:
        return str(input)

    return input


def load_cell(s, r, c):
    x = s.cell(row=r, column=c)
    x.value = convert_to_str(x.value)
    return x


def set_cell(s, r, c, value):
    cell = s.cell(row=r, column=c)
    cell.value = value


sheet = wb_obj['成長股']
stock_growth = []
for i in range(1, sheet.max_row + 1):
    stock_growth.append(load_cell(sheet, i, 1).value)

sheet = wb_obj['有發行股期的']
stock_future = []
for i in range(1, sheet.max_row + 1):
    stock_future.append(load_cell(sheet, i, 1).value)
print(stock_future)

sheet = wb_obj['50指數成分']
stock_50 = []
for i in range(1, sheet.max_row + 1):
    stock_50.append(load_cell(sheet, i, 1).value)
print(stock_50)

sheet = wb_obj['100指數成分']
stock_100 = []
for i in range(1, sheet.max_row + 1):
    stock_100.append(load_cell(sheet, i, 1).value)
print(stock_100)

sheet = wb_obj['名稱代號']
for i in range(2, sheet.max_row + 1):
    symbol = load_cell(sheet, i, 1)

    if symbol.value in stock_100:
        set_cell(sheet, i, 5, "100指數成分")

    if symbol.value in stock_50:
        set_cell(sheet, i, 5, "50指數成分")

    if symbol.value in stock_future:
        set_cell(sheet, i, 6, "有股期")

    if symbol.value in stock_growth:
        set_cell(sheet, i, 7, "成長股")


wb_obj.save(my_work_book)
