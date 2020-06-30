import sys
import traceback
import json
import datetime
import time
import random

import openpyxl
import requests

my_work_book = 'growth.xlsx'

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)

def calculate_percentage(buy, sell):
    return (sell-buy) / buy

def convert_to_str(input):
    if type(input) is str:
        return input.rstrip(' ').lstrip(' ').rstrip(' ')

    if type(input) is int:
        return str(input)

    return input


def load_cell(s, r, c):
    x = s.cell(row=r, column=c)
    x.value = convert_to_str(x.value)
    return x


def set_cell(s, r, c, value):
    cell = s.cell(row=r, column=c)
    cell.value = value


sheet = wb_obj['分析']

try:
    column = ord('U') - ord('A') + 1
    skip = []
    for i in range(62, sheet.max_row + 1):

        if i in skip:
            continue

        count = 0
        if load_cell(sheet, i, 3).value == '投信買籌多':
            for j in range(0, 1000):
                if load_cell(sheet, i + j, 3).value == '投信買籌多':
                    count = count + 1
                else:
                    break
            r = random.sample(range(count), count)
            print(r)
            for j in range(0, count):
                set_cell(sheet, i + j, column, str(r[j]))
                set_cell(sheet, i + j, column + 1, j)
            skip = range(i + 1, i + 1 + count)

except Exception as e:
    #    print(e)
    error_class = e.__class__.__name__  # 取得錯誤類型
    detail = e.args[0]  # 取得詳細內容
    cl, exc, tb = sys.exc_info()  # 取得Call Stack
    lastCallStack = traceback.extract_tb(tb)[-1]  # 取得Call Stack的最後一筆資料
    fileName = lastCallStack[0]  # 取得發生的檔案名稱
    lineNum = lastCallStack[1]  # 取得發生的行號
    funcName = lastCallStack[2]  # 取得發生的函數名稱
    errMsg = "File \"{}\", line {}, in {}: [{}] {}".format(fileName, lineNum, funcName, error_class, detail)
    print(errMsg)

wb_obj.save(my_work_book)

