import sys
import traceback
import json
import datetime
import time
import random

import openpyxl
import requests

my_work_book = 'growth.xlsx'

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)


def calculate_percentage(buy, sell):
    return (sell - buy) / buy


def convert_to_str(input):
    if type(input) is str:
        return input.rstrip(' ').lstrip(' ').rstrip(' ')

    if type(input) is int:
        return str(input)

    return input


def load_cell(s, r, c):
    x = s.cell(row=r, column=c)
    x.value = convert_to_str(x.value)
    return x


def load_value(s, r, c):
    return load_cell(s, r, c).value


def set_cell(s, r, c, value):
    cell = s.cell(row=r, column=c)
    cell.value = value


def column_idx(char):
    return ord(char.upper()) - ord('A') + 1


iteration_times = 1000
number_candidates = 2


history_gain = []

sheet = wb_obj['分析']
history_trade = []


column = ord('U') - ord('A') + 1
skip = []

day_index = {}
day_idx = 0


for i in range(62, sheet.max_row + 1):

    if i in skip:
        continue

    count = 0
    if load_cell(sheet, i, 3).value == '投信買籌多':

        date = load_value(sheet, i, column_idx('B'))

        date = load_value(sheet, i, column_idx('B'))
        symbol = load_value(sheet, i, column_idx('D'))
        name = load_value(sheet, i, column_idx('E'))
        base_price = load_value(sheet, i, column_idx('K'))
        percentage = load_value(sheet, i, column_idx('Q'))  # 4 day
        rand_idx = load_value(sheet, i, column_idx('U'))

        history_trade.append((date, symbol, name, base_price, percentage, rand_idx))

        if date not in day_index.keys():
            day_index[date] = day_idx
            day_idx = day_idx + 1


for d in day_index:
    print(d, day_index[d])


sheet = wb_obj['模擬交易']

set_cell(sheet, 3, 1, '第幾次模擬')
set_cell(sheet, 3, 2, '總獲利')
set_cell(sheet, 3, 3, '最大資金')
set_cell(sheet, 3, 4, '勝率')
set_cell(sheet, 3, 5, '平均報酬率')

columns = 8

number_trade = 1

for i in range(0, iteration_times):

    buckets = [0] * (len(day_index) + 5)
    total_gain = 0
    r = random.sample(range(10), 10)
    candidates = [str(x) for x in r[0:number_candidates]]

    rows = 4

    total_trades = 0
    total_wins = 0
    last_date = None
    total_percentage = 0

    for x in history_trade:
        #print(x)
        (date, symbol, name, base_price, percentage, rand_idx) = x
        if percentage is None:
            continue
        if rand_idx not in candidates:
            continue

        if float(base_price) > 600:
            continue

        if last_date != date:
            last_date = date
            r = random.sample(range(10), 10)
            candidates = [str(x) for x in r[0:number_candidates]]

        total_trades = total_trades + 1
        if float(percentage) > 0:
            total_wins = total_wins + 1

        total_percentage = total_percentage + float(percentage)

        day_idx = day_index[date]
        buckets[day_idx+0] = buckets[day_idx+0] + float(base_price)
        buckets[day_idx+1] = buckets[day_idx+1] + float(base_price)
        buckets[day_idx+2] = buckets[day_idx+2] + float(base_price)
        buckets[day_idx+3] = buckets[day_idx+3] + float(base_price)

        gain = float(base_price)*float(percentage)
        total_gain = total_gain + gain

        #print(date, symbol, name, base_price, gain)

        set_cell(sheet, rows, columns+0, date)
        set_cell(sheet, rows, columns+1, symbol)
        set_cell(sheet, rows, columns+2, name)
        set_cell(sheet, rows, columns+3, base_price)
        set_cell(sheet, rows, columns+4, gain)
        set_cell(sheet, rows, columns+5, percentage)
        set_cell(sheet, rows, columns+6, buckets[day_index[date]])

        rows = rows + 1

    set_cell(sheet, 2, columns+1, '第幾次模擬')
    set_cell(sheet, 3, columns+1, number_trade)

    set_cell(sheet, 2, columns+3, '最大資金')
    set_cell(sheet, 3, columns+3, max(buckets))

    set_cell(sheet, 2, columns+4, '總收益')
    set_cell(sheet, 3, columns+4, total_gain)

    set_cell(sheet, 2, columns+5, '平均報酬')
    set_cell(sheet, 3, columns+5, total_percentage * 100 / total_trades)
    #print('Total: %.2f' % total_gain)

    columns = columns + 8

    set_cell(sheet, i+4, 1, number_trade)
    set_cell(sheet, i+4, 2, total_gain)
    set_cell(sheet, i+4, 3, max(buckets))
    set_cell(sheet, i+4, 4, total_wins * 100 / total_trades)
    set_cell(sheet, i+4, 5, total_percentage * 100 / total_trades)
    history_gain.append(total_gain)

    number_trade = number_trade + 1

wb_obj.save(my_work_book)

for x in history_gain:
    print(x)


