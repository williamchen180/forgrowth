#
#
# 根據Excel裡面的標的去鉅亨網抓歷史股價，然後將之後的報酬算出來
#
#

import sys
import traceback
import json
import datetime
import time

import openpyxl
import requests

my_work_book = 'growth.xlsx'

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)

def calculate_percentage(buy, sell):
    return (sell-buy) / buy

def convert_to_str(input):
    if type(input) is str:
        return input.rstrip(' ').lstrip(' ').rstrip(' ')

    if type(input) is int:
        return str(input)

    return input


def load_cell(s, r, c):
    x = s.cell(row=r, column=c)
    x.value = convert_to_str(x.value)
    return x


def set_cell(s, r, c, value):
    cell = s.cell(row=r, column=c)
    cell.value = value


sheet = wb_obj['分析']



unix_today = datetime.date.today().strftime("%s")


try:
    for i in range(62, sheet.max_row + 1):
        date = load_cell(sheet, i, 2)
        unix_sod = date.value.strftime("%s")

        symbol = load_cell(sheet, i, 4).value
        url = 'https://ws.api.cnyes.com/charting/api/v1/history?resolution=D&symbol=TWS:%s:STOCK&from=%s&to=%s&quote=1' % (symbol, unix_today, unix_sod)

        print(url)

        r = requests.get(url)

        #print(r.content)

        data = json.loads(r.content)

        if data['statusCode'] != 200:
            print("HTTP request error!")
            continue

        history = []
        for t in data['data']['t']:
            print(datetime.datetime.utcfromtimestamp(t).strftime('%Y-%m-%d'))
            pass

        T = data['data']['t']
        O = data['data']['o']
        H = data['data']['h']
        L = data['data']['l']
        C = data['data']['c']
        V = data['data']['v']

        for o, h, l, c in zip(O, H, L, C):
            history.insert(0, (o, h, l, c))

        #print('hisroty:', history)

        if len(history) == 0:
            continue

        base_price = history[0][0]

        #print('base_price:', base_price)

        mark_days = [1, 2, 3, 4, 5, 10, 15, 20, 30]

        column_idx = ord('K') - ord('A') + 1
        set_cell(sheet, i, column_idx, base_price)
        column_idx = column_idx + 1

        for j in range(0, len(history)):
            if j+1 in mark_days:
                #print(history[j][3])
                percentage = calculate_percentage(base_price, history[j][3])
                #print(percentage)

                #set_cell(sheet, i, column_idx, percentage)
                set_cell(sheet, i, column_idx, history[j][3])

                column_idx = column_idx + 1

except Exception as e:
    #    print(e)
    error_class = e.__class__.__name__  # 取得錯誤類型
    detail = e.args[0]  # 取得詳細內容
    cl, exc, tb = sys.exc_info()  # 取得Call Stack
    lastCallStack = traceback.extract_tb(tb)[-1]  # 取得Call Stack的最後一筆資料
    fileName = lastCallStack[0]  # 取得發生的檔案名稱
    lineNum = lastCallStack[1]  # 取得發生的行號
    funcName = lastCallStack[2]  # 取得發生的函數名稱
    errMsg = "File \"{}\", line {}, in {}: [{}] {}".format(fileName, lineNum, funcName, error_class, detail)
    print(errMsg)

wb_obj.save(my_work_book)

