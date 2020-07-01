import csv
import openpyxl
from datetime import datetime
from libexcel import *

my_work_book = 'WarrantOverBuy.xlsx'

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)

#sheet = wb_obj['買超購']
#sheet = wb_obj['買超售']
#sheet = wb_obj['賣超購']
sheet = wb_obj['賣超售']

time_column = 'R'
begin_trade_column = 'Q'
call_percentage_column = 'O'
put_percentage_column = 'P'
start_row = 13

minute_mark = [5, 15, 30, 45]

bandit1 = 0.000425
bandit2 = 0.0015

stop_profit = 0.012
stop_lose = -0.006

for i in range(start_row + 1, sheet.max_row + 1):
    date = load_value(sheet, i, column_number('A'))
    symbol = load_value(sheet, i, column_number('B'))

    print(date, symbol)

    if symbol is None or date is None:
        continue

    csv_file = f'KLINE/K_{symbol}.csv'

    with open(csv_file) as f:
        rows = csv.reader(f)

        column = column_number(time_column)
        day_found = False
        base_price = 0
        call_out = False
        put_out = False
        for r in rows:
            d = datetime.strptime(r[0], '%Y/%m/%d %H:%M')

            if day_found is True and (d.year, d.month, d.day) != (date.year, date.month, date.day):
                print('finished')
                break

            if day_found is not True and (d.year, d.month, d.day) == (date.year, date.month, date.day):
                print('day found')
                day_found = True

            if day_found is True:

                if d.hour == 9 and d.minute == 5:
                    base_price = float(r[4])
                    set_cell(sheet, i, column_number(begin_trade_column), base_price)

                if base_price != 0:
                    current_price = float(r[4])
                    current_call_profit_rate = (-base_price * (1.0 + bandit1) + current_price * (
                                1.0 - bandit1 - bandit2)) / base_price
                    current_put_profit_rate = (base_price * (1 - bandit1) - current_price * (
                                1.0 + bandit1 + bandit2)) / base_price

                    if call_out is not True:
                        set_cell(sheet, i, column_number(call_percentage_column), current_call_profit_rate)
                    if current_call_profit_rate >= stop_profit or current_call_profit_rate <= stop_lose:
                        call_out = True

                    if put_out is not True:
                        set_cell(sheet, i, column_number(put_percentage_column), current_put_profit_rate)
                    if current_put_profit_rate >= stop_profit or current_put_profit_rate <= stop_lose:
                        put_out = True

                if d.minute in minute_mark:
                    set_cell(sheet, start_row, column, f'{d.hour}:{d.minute}')
                    set_cell(sheet, i, column, r[4])
                    column = column + 1

wb_obj.save(my_work_book)
