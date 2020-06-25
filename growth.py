
import openpyxl
my_work_book = 'growth.xlsx'

stock_future = {}
stock_growth = []

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)

sheet = wb_obj['有發行股期的']

def convert_to_str(input):
    if type(input) is str:
        return input.rstrip(' ').lstrip(' ').rstrip(' ')

    if type(input) is int:
        return str(input)

    return input


for i in range(1, sheet.max_row + 1):
    idx_cell = sheet.cell(row=i, column=1)
    idx_cell.value = convert_to_str(idx_cell.value)

    name_cell = sheet.cell(row=i, column=2)
    name_cell.value = convert_to_str(name_cell.value)

    market_cell = sheet.cell(row=i, column=3)
    market_cell.value = convert_to_str(market_cell.value)

    stock_future[idx_cell.value] = (name_cell.value, market_cell.value)
    # print(idx_cell.value, name_cell.value, market_cell.value)

sheet = wb_obj['名稱代號']

print(stock_future.keys())

for i in range(1, sheet.max_row + 1):
    symbol_cell = sheet.cell(row=i, column=1)
    symbol_cell.value = convert_to_str(symbol_cell.value)

    name_cell = sheet.cell(row=i, column=2)
    name_cell.value = convert_to_str(name_cell.value)

    attr_cell = sheet.cell(row=i, column=3)
    attr_cell.value = convert_to_str(attr_cell.value)

    future_cell = sheet.cell(row=i, column=4)

    print(symbol_cell.value + '|' + name_cell.value)

    if '0' <= symbol_cell.value[0] <= '9':
        if symbol_cell.value in stock_future.keys():
            print(symbol_cell.value, name_cell.value,  ' 有股期')
            future_cell.value = '有股期'

    else:
        print(symbol_cell.value, name_cell.value)


wb_obj.save(my_work_book)

exit(0)


sheet = wb_obj['成長股']

for i in range(1, sheet.max_row + 1):
    cell = sheet.cell(row=i, column=1)
    stock_growth.append(cell.value)

print(stock_growth)
