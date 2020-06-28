#
# 準備「分析」裡面的資料
#
# 補齊 日期 策略
#
# 前兩個欄位而已
#


import openpyxl
from libexcel import *

my_work_book = 'growth3.xlsx'

wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)


sheet = wb_obj['原始資料']

last_date = ''
last_policy = ''

for i in range(2, sheet.max_row + 1):
    date = load_cell(sheet, i, 1)
    policy = load_cell(sheet, i, 2)

    if date.value is not None:
        last_date = date
    else:
        set_cell(sheet, i, 1, last_date.value)

    if policy.value is not None:
        last_policy = policy
    else:
        set_cell(sheet, i, 2, last_policy.value)

wb_obj.save(my_work_book)

