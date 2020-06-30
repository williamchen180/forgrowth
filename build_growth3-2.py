#
# 準備「分析」裡面的資料
#
# 目前標頭為
#
# 日期	策略	代號	名稱	成交價	  漲跌    幅(%)  	成交量	  月線   斜率	  上通   斜率	  下通   斜率	帶寬	位階	主力連買	外資連買	投信連買	主1	主5	主10	外資%	投信%	自避%	當沖%	融資%	融券%	券資比%	乖離年線%	融資維持率	當沖損益(千)
#
# 補上
#
# 開盤	1 day	2 day	3 day	4 day	5 day	10 day	15 day	20 day	30 days
#


import os
import sys
import traceback
import json
import requests
import datetime
import openpyxl
import pickle
from libexcel import *


def calculate_percentage(buy, sell):
    return (sell - buy) / buy


def column_number(s):
    ret = 0
    s = s.upper()
    for i in range(0, len(s)):
        ret = ret * 26 + ord(s[i]) - ord('A') + 1
    return ret



my_work_book = 'growth3.xlsx'
wb_obj = openpyxl.load_workbook(my_work_book)

print(wb_obj.sheetnames)
sheet = wb_obj['分析']

unix_today = datetime.date.today().strftime("%s")
mark_days = [1, 2, 3, 4, 5, 10, 15, 20, 30]

for i in range(2, sheet.max_row + 1):
    date: datetime.datetime = load_cell(sheet, i, 1).value
    unix_sod = date.strftime("%s")
    symbol = load_cell(sheet, i, 3).value

    print('Processing ' + symbol)

    history = get_history(symbol, unix_sod, unix_today)

    if history is None:
        print(f'Cannot get {symbol} history data')
        continue

    if len(history) == 0:
        continue

    #print(f'date: {date} {unix_sod}')


    # h: {T, O, H, L, C, V}
    number_day = 0
    column_idx = 0
    base_price = -1
    sod_found = False

    for h in history:
        #print(h)
        hdate = datetime.datetime.fromtimestamp(h[0])

        if (hdate.year, hdate.month, hdate.day) == (date.year, date.month, date.day):
            print('found the date:', date)
            sod_found = True
            continue

        if sod_found is not True:
            continue

        # The base price is the Open of 'NEXT DAY'
        if base_price == -1:
            base_price = float(h[1])
            column_idx = column_number('AD')
            set_cell(sheet, i, column_idx, base_price)
            column_idx = column_idx + 1
            number_day = 1

        # print('base_price:', base_price)

        if number_day in mark_days:
            percentage = calculate_percentage(base_price, h[4])
            #set_cell(sheet, i, column_idx, percentage)
            set_cell(sheet, i, column_idx, h[4])
            column_idx = column_idx + 1

        number_day = number_day + 1



wb_obj.save(my_work_book)

